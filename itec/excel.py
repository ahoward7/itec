from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from typing import List, Dict, Tuple
import os


def exportToExcel(items: List[Dict], serializer, file_name: str = 'data.xlsx', directory: str = None) -> str:
    """Export a list of dictionaries to an Excel file"""
    if not isinstance(items, list) or not all(isinstance(item, dict) for item in items):
        raise ValueError('Input must be a list of dictionaries')

    # Get the fields from the meta class of the serializer
    fields = getFields(items, serializer)

    # Create the workbook at the file path
    wb, file_path = createWorkbook(fields, file_name, directory)

    # Apply formatting to the cells
    formatCells(wb)

    # Save the workbook
    wb.save(file_path)

    return file_path


def getFields(items: List[Dict], serializer) -> List[Dict]:
    """Get the fields from the serializer and filter the items"""
    fields = serializer.Meta.fields
    filtered_items = []
    for item in items:
        filtered_item = {field: item[field] for field in fields if field in item}
        for key, value in filtered_item.items():
            if isinstance(value, datetime):
                filtered_item[key] = value.replace(tzinfo=None)
        filtered_items.append(filtered_item)
    return filtered_items


def createWorkbook(fields: List[Dict], file_name: str = 'data.xlsx', directory: str = None) -> Tuple[Workbook, str]:
    """Create a workbook from a list of dictionaries"""
    if directory:
        if not os.path.exists(directory):
            os.makedirs(directory)
        file_path = os.path.join(directory, file_name)
    else:
        file_path = file_name

    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Extract the inventory headers
    headers = list(fields[0].keys())

    # Add the main title cell
    ws.append([None] * len(headers))
    ws.append([None] * len(headers))
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = 'Transaction Sheet'
    title_cell.font = Font(bold=True, size=20, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='5e0f1f', end_color='5e0f1f', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add another row for the losing and gaining accounts
    ws.append([None] * len(headers))
    ws.append([None] * len(headers))
    ws.merge_cells(start_row=3, start_column=1, end_row=4, end_column=len(headers) // 2)
    ws.merge_cells(start_row=3, start_column=len(headers) // 2 + 1, end_row=4, end_column=len(headers))
    losing_cell = ws.cell(row=3, column=1)
    losing_cell.value = 'Losing Account'
    losing_cell.font = Font(bold=True, size=14, color='FFFFFF')
    losing_cell.fill = PatternFill(start_color='864b57', end_color='864b57', fill_type='solid')
    losing_cell.alignment = Alignment(horizontal='center', vertical='center')
    gaining_cell = ws.cell(row=3, column=len(headers) // 2 + 1)
    gaining_cell.value = 'Gaining Account'
    gaining_cell.font = Font(bold=True, size=14, color='FFFFFF')
    gaining_cell.fill = PatternFill(start_color='864b57', end_color='864b57', fill_type='solid')
    gaining_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add the inventory headers
    ws.append(headers)

    # Add the rest of the data
    for row_data in fields:
        ws.append(list(row_data.values()))

    wb.save(file_path)

    return wb, file_path


def formatCells(wb: Workbook) -> None:
    """Format the cells in the excel workbook"""
    ws = wb.active

    header_font = Font(bold=True, color='0f5e4e')

    # Styling for the header row
    for cell in ws[5]:
        cell.value = cell.value.replace('_', ' ').title()
        cell.font = header_font

    return
